"""
Batch Link Reader for SocMed Downloader
Supports: TXT, CSV, JSON, Excel (.xlsx), Word (.docx)
"""

import json
import csv
import re
from pathlib import Path

def read_txt_file(file_path):
    """
    Read links from TXT file (1 link per line)
    
    Format:
        https://youtube.com/watch?v=...
        https://tiktok.com/@user/video/...
        # Comments are ignored
    """
    links = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                # Skip empty lines and comments
                if line and not line.startswith('#') and is_valid_url(line):
                    links.append({'url': line, 'quality': 'best', 'format': 'video'})
    except Exception as e:
        raise Exception(f"Error reading TXT file: {str(e)}")
    
    return links

def read_csv_file(file_path):
    """
    Read links from CSV file with optional metadata
    
    Format:
        url,quality,format
        https://youtube.com/...,1080,video
        https://tiktok.com/...,best,video
        https://youtube.com/...,192,audio
    
    Or simple format (just URLs):
        https://youtube.com/...
        https://tiktok.com/...
    """
    links = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            # Try to detect if file has header
            sample = f.read(1024)
            f.seek(0)
            
            has_header = 'url' in sample.lower() or 'link' in sample.lower()
            
            reader = csv.DictReader(f) if has_header else csv.reader(f)
            
            for row in reader:
                if isinstance(row, dict):
                    # CSV with headers
                    url = row.get('url') or row.get('link') or row.get('URL') or row.get('Link')
                    quality = row.get('quality', 'best')
                    format_type = row.get('format', 'video')
                else:
                    # CSV without headers (just URLs)
                    url = row[0] if row else None
                    quality = row[1] if len(row) > 1 else 'best'
                    format_type = row[2] if len(row) > 2 else 'video'
                
                if url and is_valid_url(url):
                    links.append({
                        'url': url.strip(),
                        'quality': quality,
                        'format': format_type
                    })
    except Exception as e:
        raise Exception(f"Error reading CSV file: {str(e)}")
    
    return links

def read_json_file(file_path):
    """
    Read links from JSON file
    
    Format 1 (Array of objects):
        [
            {"url": "https://youtube.com/...", "quality": "1080", "format": "video"},
            {"url": "https://tiktok.com/...", "quality": "best", "format": "video"}
        ]
    
    Format 2 (Object with links array):
        {
            "links": [
                {"url": "https://youtube.com/...", "quality": "720"},
                {"url": "https://tiktok.com/..."}
            ]
        }
    
    Format 3 (Simple array of URLs):
        [
            "https://youtube.com/...",
            "https://tiktok.com/..."
        ]
    """
    links = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Handle different JSON formats
        if isinstance(data, list):
            for item in data:
                if isinstance(item, str):
                    # Simple string URL
                    if is_valid_url(item):
                        links.append({'url': item, 'quality': 'best', 'format': 'video'})
                elif isinstance(item, dict):
                    # Object with metadata
                    url = item.get('url') or item.get('link')
                    if url and is_valid_url(url):
                        links.append({
                            'url': url,
                            'quality': item.get('quality', 'best'),
                            'format': item.get('format', 'video')
                        })
        elif isinstance(data, dict):
            # Object with links array
            link_array = data.get('links') or data.get('urls') or data.get('downloads')
            if link_array:
                return read_json_links_from_array(link_array)
    except Exception as e:
        raise Exception(f"Error reading JSON file: {str(e)}")
    
    return links

def read_json_links_from_array(link_array):
    """Helper for reading links from JSON array"""
    links = []
    for item in link_array:
        if isinstance(item, str):
            if is_valid_url(item):
                links.append({'url': item, 'quality': 'best', 'format': 'video'})
        elif isinstance(item, dict):
            url = item.get('url') or item.get('link')
            if url and is_valid_url(url):
                links.append({
                    'url': url,
                    'quality': item.get('quality', 'best'),
                    'format': item.get('format', 'video')
                })
    return links

def read_excel_file(file_path):
    """
    Read links from Excel file (.xlsx)
    
    Format:
        Column A: URL/Link
        Column B: Quality (optional)
        Column C: Format (optional)
    
    First row can be header or data
    """
    links = []
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # Check if first row is header
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        has_header = any(str(cell).lower() in ['url', 'link', 'quality', 'format'] 
                        for cell in first_row if cell)
        
        start_row = 2 if has_header else 1
        
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if not row or not row[0]:
                continue
            
            url = str(row[0]).strip()
            quality = str(row[1]).strip() if len(row) > 1 and row[1] else 'best'
            format_type = str(row[2]).strip() if len(row) > 2 and row[2] else 'video'
            
            if is_valid_url(url):
                links.append({
                    'url': url,
                    'quality': quality,
                    'format': format_type
                })
        
        wb.close()
    except ImportError:
        raise Exception("openpyxl not installed. Install: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")
    
    return links

def read_word_file(file_path):
    """
    Read links from Word file (.docx)
    Extracts all URLs from paragraphs and hyperlinks
    """
    links = []
    try:
        from docx import Document
        
        doc = Document(file_path)
        
        # Extract from paragraphs
        for para in doc.paragraphs:
            text = para.text.strip()
            # Find URLs in text
            urls = extract_urls_from_text(text)
            for url in urls:
                if is_valid_url(url):
                    links.append({'url': url, 'quality': 'best', 'format': 'video'})
        
        # Extract from hyperlinks
        for rel in doc.part.rels.values():
            if "hyperlink" in rel.target_ref:
                url = rel.target_ref
                if is_valid_url(url):
                    links.append({'url': url, 'quality': 'best', 'format': 'video'})
    
    except ImportError:
        raise Exception("python-docx not installed. Install: pip install python-docx")
    except Exception as e:
        raise Exception(f"Error reading Word file: {str(e)}")
    
    return links

def extract_urls_from_text(text):
    """Extract URLs from text using regex"""
    url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
    urls = re.findall(url_pattern, text)
    return urls

def is_valid_url(url):
    """Check if string is a valid URL"""
    if not url or not isinstance(url, str):
        return False
    
    url = url.strip()
    
    # Basic URL validation
    if not url.startswith(('http://', 'https://')):
        return False
    
    # Check for supported platforms (optional, can be expanded)
    supported_domains = [
        'youtube.com', 'youtu.be',  # YouTube
        'tiktok.com',                # TikTok
        'instagram.com',             # Instagram
        'facebook.com', 'fb.watch',  # Facebook
        'twitter.com', 'x.com',      # Twitter/X
    ]
    
    # Allow any URL (yt-dlp will validate)
    # But check minimum length and structure
    if len(url) < 10 or ' ' in url:
        return False
    
    return True

def read_batch_file(file_path):
    """
    Auto-detect file format and read links
    
    Supported formats:
    - .txt: Plain text, 1 link per line
    - .csv: CSV with optional headers (url, quality, format)
    - .json: JSON array or object
    - .xlsx: Excel spreadsheet
    - .docx: Word document
    
    Returns:
        list: List of dicts with keys: url, quality, format
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    extension = file_path.suffix.lower()
    
    readers = {
        '.txt': read_txt_file,
        '.csv': read_csv_file,
        '.json': read_json_file,
        '.xlsx': read_excel_file,
        '.xls': read_excel_file,
        '.docx': read_word_file,
        '.doc': read_word_file,
    }
    
    reader_func = readers.get(extension)
    
    if not reader_func:
        raise ValueError(f"Unsupported file format: {extension}")
    
    links = reader_func(file_path)
    
    # Remove duplicates while preserving order
    seen = set()
    unique_links = []
    for link in links:
        url = link['url']
        if url not in seen:
            seen.add(url)
            unique_links.append(link)
    
    return unique_links

# Example usage and testing
if __name__ == "__main__":
    # Test TXT format
    print("Testing batch_reader.py...")
    
    # Create sample TXT file
    sample_txt = "test_links.txt"
    with open(sample_txt, 'w') as f:
        f.write("https://www.youtube.com/watch?v=dQw4w9WgXcQ\n")
        f.write("https://www.tiktok.com/@test/video/123456\n")
        f.write("# This is a comment\n")
        f.write("https://www.instagram.com/reel/ABC123/\n")
    
    try:
        links = read_batch_file(sample_txt)
        print(f"Found {len(links)} links:")
        for i, link in enumerate(links, 1):
            print(f"  {i}. {link['url']}")
        
        import os
        os.remove(sample_txt)
        print("\nTest passed!")
    except Exception as e:
        print(f"Error: {e}")
